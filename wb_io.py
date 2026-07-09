"""
wb_io.py — one robust entry point for opening user workbooks.

Motivation (WORKORDER_ingestion_robustness.md): real underwriting models arrive
with tens of thousands of junk `definedName` entries (residue of copying sheets
between workbooks), most rotted to `#REF!`, plus orphaned external links. A full
(`read_only=False`) `openpyxl.load_workbook` builds that whole object graph and
pays ~9× the cost of a `read_only` load. Stacked across a pipeline that loads the
same file many times, that blows the Cloud Run request timeout and the browser
shows the opaque "Load failed."

This module gives every call site ONE loader with three guarantees:

  1. read_only by default — the vast majority of callers only read cell values,
     and `read_only` skips the defined-name / external-link graph entirely.
  2. A wall-clock timeout — a pathological file raises a clear, catchable
     `WorkbookLoadError` instead of hanging the request until Cloud Run severs it.
  3. Full loads are sanitized — when a caller genuinely needs the mutable object
     model (`read_only=False`), we first strip `#REF!` defined names and orphaned
     external links onto an in-memory temp COPY (never the user's file) so the full
     load stops paying for junk.

Optionally, a request-scoped cache (`workbook_cache()`) ensures a given file is
loaded at most once per (mode) per request — callers keep calling `wb.close()`;
for a cached workbook that call is neutered and the cache closes it at teardown.
"""
from __future__ import annotations

import contextlib
import contextvars
import logging
import os
import re
import sys
import tempfile
import threading
import zipfile
from pathlib import Path
from typing import Any, Iterator

import openpyxl

log = logging.getLogger("fb.wb_io")
if not log.handlers:
    _h = logging.StreamHandler(sys.stdout)
    _h.setFormatter(logging.Formatter("[fb.wb_io] %(asctime)s %(levelname)s %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

# Wall-clock ceiling for a single load. A healthy 15MB workbook (the upload cap)
# loads full in a few seconds; this is a safety net against a truly pathological
# file, not a tuning knob for the common case.
DEFAULT_TIMEOUT = float(os.getenv("WB_LOAD_TIMEOUT", "25"))


class WorkbookLoadError(Exception):
    """A workbook could not be opened within the guarantees of safe_load_workbook.

    Always carries a short, user-safe `reason`. Callers catch this and DEGRADE
    (skip an enrichment, return `limited` mode) rather than letting an opaque
    error sever the request.
    """

    def __init__(self, reason: str):
        super().__init__(reason)
        self.reason = reason


# ---------------------------------------------------------------------------
# Request-scoped cache (opt-in)
# ---------------------------------------------------------------------------

_cache_var: contextvars.ContextVar[dict | None] = contextvars.ContextVar(
    "wb_io_cache", default=None
)


@contextlib.contextmanager
def workbook_cache() -> Iterator[None]:
    """Within this context, safe_load_workbook loads each (path, mode) at most once.

    Callers keep calling `wb.close()`; for a cached workbook that is neutered to a
    no-op and the real close runs here at teardown. Wrap the whole /api/analyze
    request in this so the pipeline's many loads collapse to one-per-mode.
    """
    token = _cache_var.set({})
    try:
        yield
    finally:
        cache = _cache_var.get() or {}
        for wb in cache.values():
            closer = getattr(wb, "_wb_io_real_close", None)
            with contextlib.suppress(Exception):
                (closer or wb.close)()
        _cache_var.reset(token)


def _neuter_close(wb: Any) -> Any:
    """Make wb.close() a no-op so a caller can't close a cached, shared workbook.

    The request cache holds the real closer and calls it at teardown. If the
    workbook object won't accept the attribute (e.g. __slots__), we quietly skip
    caching-safety for it — worst case it gets closed early and re-loaded.
    """
    try:
        wb._wb_io_real_close = wb.close
        wb.close = lambda: None  # type: ignore[method-assign]
    except Exception:
        pass
    return wb


# ---------------------------------------------------------------------------
# Sanitize pathological workbooks (full loads only)
# ---------------------------------------------------------------------------

_DEFINED_NAME_RE = re.compile(r"<definedName\b[^>]*>.*?</definedName>", re.DOTALL)
_EXTERNAL_REFS_RE = re.compile(r"<externalReferences\b.*?</externalReferences>", re.DOTALL)


def _sanitize_to_temp(src: Path) -> str | None:
    """Write a cleaned COPY of `src` and return its path, or None if nothing to do.

    Drops `<definedName>` entries whose formula is `#REF!` (dead) and, if present,
    the `<externalReferences>` block (orphaned links Excel would try to resolve).
    Everything else — every sheet, value, formula, style — is copied byte-for-byte.
    Never mutates the user's original file. On any error, returns None so the
    caller falls back to loading the original.
    """
    try:
        with zipfile.ZipFile(src) as zin:
            try:
                wbxml = zin.read("xl/workbook.xml").decode("utf-8")
            except KeyError:
                return None  # not a normal xlsx; let the real loader speak

            dead = 0

            def _drop_ref(m: re.Match) -> str:
                nonlocal dead
                if "#REF!" in m.group(0):
                    dead += 1
                    return ""
                return m.group(0)

            cleaned = _DEFINED_NAME_RE.sub(_drop_ref, wbxml)
            ext_stripped = bool(_EXTERNAL_REFS_RE.search(cleaned))
            if ext_stripped:
                cleaned = _EXTERNAL_REFS_RE.sub("", cleaned)

            if dead == 0 and not ext_stripped:
                return None  # healthy workbook — don't pay to rewrite it

            fd, tmp = tempfile.mkstemp(suffix=".xlsx", prefix="wbio_")
            os.close(fd)
            with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as zout:
                for item in zin.infolist():
                    data = zin.read(item.filename)
                    if item.filename == "xl/workbook.xml":
                        data = cleaned.encode("utf-8")
                    zout.writestr(item, data)
        log.info("sanitized %s — dropped %d #REF! name(s)%s", src.name, dead,
                 ", stripped external refs" if ext_stripped else "")
        return tmp
    except Exception as e:  # pragma: no cover - defensive
        log.warning("sanitize skipped for %s (%s: %s)", src.name, type(e).__name__, e)
        return None


# ---------------------------------------------------------------------------
# The loader
# ---------------------------------------------------------------------------

def _run_with_timeout(fn, timeout: float, what: str):
    """Run fn() on a worker thread; raise WorkbookLoadError if it overruns.

    openpyxl offers no cooperative cancellation, so on breach the worker is left
    to finish (daemon) and we raise. That's acceptable: the point is to free the
    REQUEST, not the CPU, before Cloud Run severs it.
    """
    result: dict[str, Any] = {}

    def _target():
        try:
            result["value"] = fn()
        except BaseException as e:  # noqa: BLE001 - propagate to caller thread
            result["error"] = e

    t = threading.Thread(target=_target, daemon=True)
    t.start()
    t.join(timeout)
    if t.is_alive():
        raise WorkbookLoadError(
            f"{what} timed out after {timeout:.0f}s — the workbook is unusually heavy."
        )
    if "error" in result:
        e = result["error"]
        raise WorkbookLoadError(f"{what} failed ({type(e).__name__}: {e})")
    return result["value"]


def safe_load_workbook(
    path: str | Path,
    *,
    data_only: bool = True,
    read_only: bool = True,
    need_formulas: bool = False,
    timeout: float = DEFAULT_TIMEOUT,
):
    """Open a workbook robustly. THE single entry point — route all loads here.

    Args:
      data_only:     read cached values (True) or formula strings (False).
      read_only:     default True — skips the defined-name / external-link graph.
                     Pass False ONLY when you need the mutable object model
                     (`wb.defined_names`, styles, insert/delete); document why at
                     the call site. Full loads are sanitized first (see module doc).
      need_formulas: convenience alias forcing data_only=False (formula strings).
                     Available in read_only mode too.
      timeout:       wall-clock ceiling; breach raises WorkbookLoadError.

    Returns an openpyxl workbook. Raises WorkbookLoadError (catchable, user-safe)
    on any failure — callers must degrade, never let it propagate opaquely.
    """
    path = Path(path)
    if need_formulas:
        data_only = False
    key = (str(path.resolve()), data_only, read_only)

    cache = _cache_var.get()
    if cache is not None and key in cache:
        return cache[key]

    if not path.exists():
        raise WorkbookLoadError(f"file not found: {path.name}")

    def _load():
        # read_only loads are cheap and skip the junk graph — no sanitize needed.
        if read_only:
            return openpyxl.load_workbook(path, data_only=data_only, read_only=True)
        # Full load: strip dead names / orphaned external links onto a temp copy
        # first so the object-graph build stops paying for junk. Delete the temp
        # right after — a full load reads the whole archive into memory at open.
        tmp = _sanitize_to_temp(path)
        target = tmp or str(path)
        try:
            return openpyxl.load_workbook(target, data_only=data_only, read_only=False)
        finally:
            if tmp:
                with contextlib.suppress(OSError):
                    os.remove(tmp)

    mode = f"{'read_only' if read_only else 'full'}/{'values' if data_only else 'formulas'}"
    wb = _run_with_timeout(_load, timeout, f"loading {path.name} ({mode})")

    if cache is not None:
        cache[key] = _neuter_close(wb)
    return wb
