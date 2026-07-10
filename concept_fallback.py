"""
concept_fallback.py — GPT-as-reader gap-fill for the deal_truth concept model.

Runs only when the deterministic pipeline (workbook_map + deal_truth) found
ZERO usable candidates for a static/labeled-assumption concept anywhere in the
workbook. For such a concept, sends a tight per-concept GPT call against the
top-priority sheet(s) for that concept's block kind and either returns a
canonical-shaped record (still subject to the caller's own domain-plausibility
check) or None if not found.

Scope is deliberately narrow: static, single-cell "read a labeled assumption"
concepts only (purchase_price, total_cost, debt, equity, ltc, ltv, exit_cap,
exit_value, interest_rate, debt_yield, dscr). Operating-trajectory concepts
(noi/revenue/opex/capex/debt_service) and cash-flow-validated returns are
intentionally out of scope — see deal_truth.py's _GAP_FILL_CONCEPTS comment.

Cheap by design: one GPT call per missing concept, each scoped to at most two
sheets rendered as a compact text grid (~1-3K tokens). Uses MODEL_FAST and is
silently skipped when no OpenAI key is set.
"""
from __future__ import annotations
import json
import logging
import sys
from pathlib import Path
from typing import Any

from scenarios._llm import client, MODEL_FAST, REASONING_EFFORT, llm_available
from workbook_map import _CONCEPT_VOCAB, _BLOCK_FOR_CONCEPT, _RATE, _MULTIPLE, _MONEY
from flexible_extractor import sorted_sheets_by_priority
from metric_resolver import parse_numeric_value
from wb_io import safe_load_workbook

log = logging.getLogger("fb.conceptfallback")
if not log.handlers:
    h = logging.StreamHandler(sys.stdout)
    h.setFormatter(logging.Formatter("[fb.conceptfallback] %(asctime)s %(levelname)s %(message)s"))
    log.addHandler(h)
    log.setLevel(logging.INFO)

CONCEPT_FALLBACK_VERSION = "concept_fallback.v1"

_MAX_ROWS_PER_SHEET = 80
_MAX_COLS_PER_SHEET = 20
_MAX_TARGET_SHEETS = 2

_VOCAB_BY_CONCEPT = dict(_CONCEPT_VOCAB)


def _domain_hint(concept: str) -> str:
    if concept in _RATE:
        return "a rate/percentage — return as a fraction (0.055) or percent (5.5), whichever the cell shows"
    if concept in _MULTIPLE:
        return "a multiple or ratio (e.g. 1.25 for a 1.25x DSCR)"
    if concept in _MONEY:
        return "a whole-dollar amount for the deal (not a per-unit, per-SF, or percentage figure)"
    return "a number"


SYSTEM_PROMPT = """\
You are a precise data extractor for a real estate underwriting model.
You will be shown ONE sheet from an Excel workbook and asked to find ONE specific concept.

CRITICAL RULES:
- Return only what is literally in the sheet. If the value is not present, say not found.
- DO NOT SUBSTITUTE. If the row whose label matches the requested concept has an
  EMPTY value cell, the concept is NOT FOUND — never borrow a value from an
  adjacent or nearby row that is a DIFFERENT concept.
- The matched row label must be SEMANTICALLY the requested concept, not merely
  near it or numerically plausible.
- The value must come from a single cell. Cite the cell (e.g. "C11").
- If ambiguous (multiple plausible cells), pick the one whose ROW LABEL is most
  semantically specific to the requested concept. Cite your reasoning.
- Numbers must be returned as numbers.

Return ONLY JSON of this shape:
{
  "found": true,
  "value": <number>,
  "cell": "C11",
  "label_in_sheet": "Loan to Value",
  "reasoning": "Row 11 is the only row labeled LTV; value is directly to its right."
}
OR if not found:
{
  "found": false,
  "reasoning": "No row in this sheet labels LTV. Searched rows 1-80."
}
No prose outside the JSON. No code fences.
"""


def _sheet_to_text_block(file_path: Path, sheet_name: str) -> str:
    """Render a single sheet as compact text grid for GPT."""
    import openpyxl
    try:
        wb = safe_load_workbook(file_path, data_only=True, read_only=False)
    except Exception as e:
        return f"(could not load workbook: {e})"

    if sheet_name not in wb.sheetnames:
        return f"(sheet {sheet_name!r} not found in workbook)"

    ws = wb[sheet_name]
    lines: list[str] = []
    for r in range(1, min(ws.max_row, _MAX_ROWS_PER_SHEET) + 1):
        for c in range(1, min(ws.max_column, _MAX_COLS_PER_SHEET) + 1):
            v = ws.cell(row=r, column=c).value
            if v is None or str(v).strip() == "":
                continue
            cell_ref = openpyxl.utils.get_column_letter(c) + str(r)
            s = str(v)
            if len(s) > 60:
                s = s[:57] + "..."
            lines.append(f"  {cell_ref}: {s}")
    wb.close()
    return "\n".join(lines)


def _target_sheets(concept: str, m: dict) -> list[str]:
    """Pick up to _MAX_TARGET_SHEETS sheets likely to hold this concept.

    Prefers sheets whose static_blocks/timeseries_blocks carry this concept's
    block kind; falls back to generic name-priority ranking when nothing
    matches (mirrors the old metric_fallback.py's catch-all behavior)."""
    block_kind = _BLOCK_FOR_CONCEPT.get(concept)
    sheets = m.get("sheets") or {}
    matches: list[str] = []
    if block_kind:
        for name, info in sheets.items():
            if block_kind in (info.get("static_blocks") or []):
                matches.append(name)
        if not matches:
            for b in m.get("timeseries_blocks") or []:
                if b.get("kind") == block_kind and b.get("sheet") not in matches:
                    matches.append(b["sheet"])

    if not matches:
        matches = sorted_sheets_by_priority(list(sheets.keys()))

    return matches[:_MAX_TARGET_SHEETS]


def _one_concept_gpt_call(concept: str, sheet_name: str, sheet_text: str) -> dict:
    aliases = ", ".join(_VOCAB_BY_CONCEPT.get(concept, ()))
    user_msg = (
        f"CONCEPT TO FIND: {concept}\n"
        f"OFTEN LABELED AS: {aliases or '(no known aliases)'}\n"
        f"EXPECTED KIND:    {_domain_hint(concept)}\n"
        f"\n"
        f"SHEET: {sheet_name}\n"
        f"{sheet_text}\n"
    )
    try:
        response = client.chat.completions.create(
            model=MODEL_FAST,
            reasoning_effort=REASONING_EFFORT,
            messages=[
                {"role": "system", "content": SYSTEM_PROMPT},
                {"role": "user",   "content": user_msg},
            ],
        )
        raw = response.choices[0].message.content.strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            if raw.startswith("json"):
                raw = raw[4:]
        return json.loads(raw)
    except json.JSONDecodeError as e:
        log.error("Gap-fill JSON parse failed for %s: %s", concept, e)
        return {"found": False, "reasoning": f"GPT response unparseable: {e}"}
    except Exception as e:
        log.error("Gap-fill API call failed for %s: %s", concept, e)
        return {"found": False, "reasoning": f"API error: {e}"}


def find_concept(concept: str, file_path: Path, m: dict) -> dict[str, Any] | None:
    """
    Try to find this concept via a targeted GPT read of top-priority sheets.

    Returns a canonical-shaped candidate:
        {concept, value, source, display, role, provenance, conflict, method}
    or None if GPT could not find it (or no API key is configured). The
    caller is responsible for domain-plausibility gating the returned value
    before trusting it — this function does not validate the value itself.
    """
    if not llm_available():
        return None

    file_path = Path(file_path)
    target_sheets = _target_sheets(concept, m)
    if not target_sheets:
        log.info("Gap-fill SKIP for %s — no sheets available", concept)
        return None

    for sheet in target_sheets:
        sheet_text = _sheet_to_text_block(file_path, sheet)
        if not sheet_text:
            continue

        log.info("Gap-fill ATTEMPT for %s on sheet %r", concept, sheet)
        result = _one_concept_gpt_call(concept, sheet, sheet_text)

        if result.get("found"):
            value, ok = parse_numeric_value(result.get("value"))
            if not ok:
                log.info("Gap-fill FOUND for %s on %s!%s but value is non-numeric: %r",
                         concept, sheet, result.get("cell"), result.get("value"))
                continue

            log.info("Gap-fill FOUND for %s on %s!%s — %r",
                      concept, sheet, result.get("cell"), str(value)[:50])
            display = f"{sheet}!{result.get('cell')}"
            # Feed the loop: GPT found a concept the deterministic catalog MISSED —
            # the matched label is a prime candidate for a new alias-catalog entry.
            try:
                from learning_store import record_resolution, file_fingerprint, FILLED
                record_resolution(layer="concept_fallback", concept=concept, decision=FILLED,
                                   file=Path(file_path).name, file_hash=file_fingerprint(file_path),
                                   label=result.get("label_in_sheet"), chosen_value=value,
                                   chosen_cell=display, reason=(result.get("reasoning") or "")[:160])
            except Exception:  # pragma: no cover - never block extraction
                pass
            return {
                "concept":     concept,
                "value":       value,
                "source":      display,
                "display":     display,
                "role":        None,
                "provenance":  None,
                "conflict":    False,
                "method":      "gpt_fallback",
                "fallback_reasoning": result.get("reasoning"),
                "matched_label": result.get("label_in_sheet"),
            }

    log.info("Gap-fill NOT FOUND for %s after %d sheet(s)", concept, len(target_sheets))
    return None
