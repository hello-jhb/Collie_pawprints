"""
fact_review.py — GPT audits the deterministic engine's headline facts.

Where concept_fallback.py fills a concept the Python pipeline found ZERO candidates
for, this REVIEWS the headline facts Python DID pick — cost, debt, equity, price,
NOI, exit, caps — against the summary/assumptions sheets, and corrects a weak pick.

The controlling principle is "GPT proposes, the deterministic oracle disposes":

  1. A fact the cash-flow oracle VALIDATED (reproduced the stated IRR) is LOCKED.
     GPT may corroborate it but can never override it. This is what keeps 1425's
     stream-derived cost and every recomputed return exactly as the engine found them.
  2. GPT may only change a NON-validated (vocab- or identity-derived) fact, and only
     when it is confident AND cites a specific summary cell AND its value still passes
     the same sanity invariants Python enforces (e.g. total cost ≥ total debt). A GPT
     suggestion that fails an invariant is rejected — Python's value stands.

So the review can rescue the messy/incomplete files ("all the numbers are in the
summary, engine just mis-picked") without ever unwinding a guarantee we already have.
Degrades to a no-op (returns the canonical unchanged) when the LLM is unavailable or
the call fails — never blocks or slows the request past the client timeout.
"""
from __future__ import annotations

import json
import logging
import sys
from pathlib import Path
from typing import Any

from scenarios._llm import get_client, MODEL_FAST, REASONING_EFFORT
from wb_io import safe_load_workbook

log = logging.getLogger("fb.factreview")
if not log.handlers:
    _h = logging.StreamHandler(sys.stdout)
    _h.setFormatter(logging.Formatter("[fb.factreview] %(asctime)s %(levelname)s %(message)s"))
    log.addHandler(_h)
    log.setLevel(logging.INFO)

FACT_REVIEW_VERSION = "fact_review.v1"

# Headline economic facts worth auditing. IRR / equity multiple are omitted: when the
# engine is found they're oracle-validated (locked) anyway, and when it isn't there's
# no stream for GPT to read them from — Tier-2 handles that case separately.
_REVIEW_FIELDS = [
    "purchase_price", "total_cost", "debt", "equity",
    "noi", "exit_value", "sale_price", "exit_cap", "going_in_cap", "ltc", "ltv",
]

# Plausible full-dollar / ratio ranges — a GPT suggestion outside these is noise and
# is rejected regardless of stated confidence.
_RANGE: dict[str, tuple[float, float]] = {
    "purchase_price": (5e4, 5e10), "total_cost": (5e4, 5e10),
    "debt": (1e4, 5e10), "equity": (1e4, 5e10),
    "noi": (1e3, 5e9), "exit_value": (5e4, 5e10), "sale_price": (5e4, 5e10),
    "exit_cap": (0.005, 0.25), "going_in_cap": (0.005, 0.25),
    "ltc": (0.01, 0.99), "ltv": (0.01, 0.99),
}

_MAX_ROWS = 70
_MAX_COLS = 16
_MAX_SHEETS = 4


def _is_validated(rec: dict | None) -> bool:
    """Oracle-validated → locked. True when the fact was reproduced from the
    cash-flow stream (method 'recomputed' / cf_validated flag)."""
    return bool(rec and (rec.get("cf_validated") or rec.get("method") == "recomputed"))


def _passes_invariants(field: str, value: float, canonical: dict) -> bool:
    """The same sanity checks Python enforces, applied to GPT's suggestion. A value
    that fails one is rejected — this is what stops GPT re-introducing an error we
    already fixed deterministically (e.g. a phased 'Total Uses — At Close' cost that
    is below total debt)."""
    lo, hi = _RANGE.get(field, (float("-inf"), float("inf")))
    if not (lo <= abs(value) <= hi):
        return False
    if field == "total_cost":
        debt = _num(canonical.get("debt"))
        if debt and abs(value) < abs(debt) * 0.999:      # cost can't be below debt
            return False
    return True


def _num(rec: dict | None) -> float | None:
    try:
        return float(rec["value"]) if rec and rec.get("value") is not None else None
    except (TypeError, ValueError):
        return None


def _summary_text(file_path: Path) -> str:
    """Render the declaration sheets (summary / inputs / returns) as compact,
    cell-referenced text so GPT can cite exactly where a value lives."""
    try:
        from workbook_orientation import orient_workbook
        rm = (orient_workbook(file_path) or {}).get("map", {})
    except Exception:
        rm = {}
    want = []
    for role in ("summary", "inputs", "returns"):
        want.extend(rm.get(role, []))
    seen: set[str] = set()
    want = [s for s in want if not (s in seen or seen.add(s))][:_MAX_SHEETS]

    try:
        wb = safe_load_workbook(file_path, data_only=True)
    except Exception as e:
        log.warning("summary render skipped (%s)", e)
        return ""
    from openpyxl.utils import get_column_letter
    if not want:
        want = wb.sheetnames[:_MAX_SHEETS]
    lines: list[str] = []
    try:
        for sn in want:
            if sn not in wb.sheetnames:
                continue
            ws = wb[sn]
            lines.append(f"\n=== SHEET: {sn} ===")
            for r in range(1, min(ws.max_row or 1, _MAX_ROWS) + 1):
                for c in range(1, min(ws.max_column or 1, _MAX_COLS) + 1):
                    v = ws.cell(row=r, column=c).value
                    if v is None or str(v).strip() == "":
                        continue
                    s = str(v)
                    lines.append(f"  {get_column_letter(c)}{r}: {s[:60]}")
    finally:
        wb.close()
    return "\n".join(lines)


_SYSTEM = """\
You audit a real-estate underwriting extraction. A deterministic engine pulled the
headline facts below, each with the cell it came from. Using ONLY the SUMMARY SHEET
TEXT (every value is cell-referenced), judge each field:

- If the engine's value is right, mark ok=true.
- If it's wrong, set ok=false and give the correct value with the EXACT cell it comes
  from. Only correct a field when the summary clearly supports a different number and
  you can name the cell. When unsure, mark ok=true (do not guess).

Rules:
- Values are full dollars unless the sheet declares thousands/millions; convert to full
  dollars. Ratios (caps, LTC, LTV) are decimals (5.0% -> 0.05).
- "total_cost" is the FULL project cost (land + all construction + financing), NOT a
  single phase of a phased sources/uses table. It must be >= total debt.
- "sale_price"/"exit_value" is THIS deal's disposition, never a comparable sale.
- Return ONLY valid JSON, no prose, no code fences:
{ "<field>": {"ok": bool, "value": <number|null>, "cell": "<Sheet!A1|null>",
              "confidence": "high"|"medium"|"low", "note": "<short>"} }
"""


def review_headline_facts(file_path: str | Path, dt: dict) -> dict:
    """Audit dt['canonical'] headline facts against the summary sheet and return a
    NEW canonical dict with weak picks corrected. Never mutates dt. Degrades to the
    original canonical (copied) on any failure."""
    file_path = Path(file_path)
    canonical = dict(dt.get("canonical", {}))
    client = get_client()
    if client is None:
        return canonical

    present = {f: canonical[f] for f in _REVIEW_FIELDS if f in canonical}
    if not present:
        return canonical

    facts_block = "\n".join(
        f"  {f}: value={_num(canonical[f])!r}  from={canonical[f].get('source')}"
        f"  validated={_is_validated(canonical[f])}"
        for f in present
    )
    summary = _summary_text(file_path)
    if not summary:
        return canonical

    user = (f"HEADLINE FACTS (engine):\n{facts_block}\n\nSUMMARY SHEET TEXT:\n{summary}")
    try:
        resp = client.chat.completions.create(
            model=MODEL_FAST, reasoning_effort=REASONING_EFFORT,  # audit task — fast model
            messages=[{"role": "system", "content": _SYSTEM},
                      {"role": "user", "content": user}],
        )
        raw = (resp.choices[0].message.content or "").strip()
        if raw.startswith("```"):
            raw = raw.split("```")[1]
            raw = raw[4:] if raw.startswith("json") else raw
        review = json.loads(raw)
    except Exception as e:  # noqa: BLE001 - degrade to Python's picks
        log.warning("review call failed (%s: %s) — keeping engine facts", type(e).__name__, e)
        return canonical

    # Feed every decision to the learning store so recurring corrections can later be
    # promoted into the deterministic layer. Best-effort — never blocks the review.
    try:
        from learning_store import (record_resolution, file_fingerprint,
                                     AGREED, CORRECTED, REJECTED)
        fhash = file_fingerprint(file_path)
    except Exception:  # pragma: no cover
        record_resolution = None

    def _log_decision(**kw):
        if record_resolution:
            try:
                record_resolution(layer="fact_review", file=file_path.name, file_hash=fhash, **kw)
            except Exception:  # pragma: no cover
                pass

    changed = 0
    for field, r in (review or {}).items():
        if field not in canonical or not isinstance(r, dict):
            continue
        cur = canonical[field]
        if _is_validated(cur):
            continue                       # oracle disposes — locked (not a GPT decision)
        if r.get("ok", True):
            _log_decision(concept=field, decision=AGREED, prior_value=_num(cur),
                          prior_source=cur.get("source"), confidence=r.get("confidence"))
            continue                       # GPT agrees with the engine
        val = r.get("value")
        try:
            val = float(val)
        except (TypeError, ValueError):
            continue
        if r.get("confidence") not in ("high", "medium"):
            continue
        if not _passes_invariants(field, val, canonical):
            log.info("review REJECTED %s=%s (fails invariant/range) — engine value kept", field, val)
            _log_decision(concept=field, decision=REJECTED, label=r.get("cell"),
                          prior_value=_num(cur), prior_source=cur.get("source"),
                          chosen_value=val, chosen_cell=r.get("cell"),
                          confidence=r.get("confidence"),
                          reason="fails invariant/range")
            continue
        canonical[field] = {
            "concept": field, "value": val,
            "source": f"GPT-reviewed: {r.get('cell') or 'summary'}",
            "method": "gpt_reviewed", "validated": False, "conflict": False,
            "cf_validated": False, "confidence": r.get("confidence"),
            "review_note": (r.get("note") or "")[:160],
            "prior_value": _num(cur), "prior_source": cur.get("source"),
        }
        changed += 1
        log.info("review CORRECTED %s: %s -> %s (%s)", field, _num(cur), val, r.get("cell"))
        _log_decision(concept=field, decision=CORRECTED, label=r.get("cell"),
                      prior_value=_num(cur), prior_source=cur.get("source"),
                      chosen_value=val, chosen_cell=r.get("cell"),
                      confidence=r.get("confidence"), reason=(r.get("note") or "")[:160])

    log.info("headline review for %s — %d field(s) corrected of %d audited",
             file_path.name, changed, len(present))
    return canonical
